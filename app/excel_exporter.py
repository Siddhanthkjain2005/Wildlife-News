from pathlib import Path
from threading import Lock

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import settings
from app.models import NewsItem

NEWS_SHEET = "news_items"
LIVE_EVENTS_SHEET = "live_events"

NEWS_HEADERS = [
    "published_at",
    "title",
    "summary",
    "source",
    "language",
    "ai_score",
    "confidence",
    "risk_score",
    "crime_type",
    "species",
    "state",
    "district",
    "location",
    "network_indicator",
    "repeat_indicator",
    "intel_summary",
    "intel_points",
    "likely_smuggling_route",
    "enforcement_recommendation",
    "confidence_explanation",
    "duplicate_confidence",
    "source_count",
    "report_count",
    "merged_sources",
    "url",
    "ai_reason",
    "last_seen_at",
]

LIVE_EVENT_HEADERS = [
    "event_time_utc",
    "event_type",
    "severity",
    "published_at",
    "title",
    "source",
    "language",
    "ai_score",
    "risk_score",
    "crime_type",
    "species",
    "state",
    "district",
    "likely_smuggling_route",
    "enforcement_recommendation",
    "confidence_explanation",
    "source_count",
    "merged_sources",
    "duplicate_confidence",
    "provider",
    "query",
    "url",
]

excel_write_lock = Lock()


def _ensure_workbook(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    default_sheet = wb.active
    default_sheet.title = NEWS_SHEET
    wb.save(path)
    return wb


def _ensure_headers(sheet, headers: list[str]) -> None:
    if sheet.max_row < 1:
        sheet.append(headers)
        return
    existing = [sheet.cell(row=1, column=i + 1).value for i in range(len(headers))]
    if existing != headers:
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(headers)


def export_news_to_excel(db: Session) -> int:
    stmt = select(NewsItem).where(NewsItem.is_poaching.is_(True)).order_by(NewsItem.published_at.desc())
    rows = db.execute(stmt).scalars().all()
    path = Path(settings.excel_path)

    with excel_write_lock:
        wb = _ensure_workbook(path)
        news_sheet = wb[NEWS_SHEET] if NEWS_SHEET in wb.sheetnames else wb.create_sheet(NEWS_SHEET)
        _ensure_headers(news_sheet, NEWS_HEADERS)

        if news_sheet.max_row > 1:
            news_sheet.delete_rows(2, news_sheet.max_row - 1)

        for item in rows:
            news_sheet.append(
                [
                    item.published_at.isoformat(sep=" ", timespec="seconds"),
                    item.title,
                    item.summary,
                    item.source,
                    item.language,
                    round(item.ai_score, 4),
                    round(item.confidence, 4),
                    item.risk_score,
                    item.crime_type,
                    item.species,
                    item.state,
                    item.district,
                    item.location,
                    item.network_indicator,
                    item.repeat_indicator,
                    item.intel_summary,
                    item.intel_points,
                    item.likely_smuggling_route,
                    item.enforcement_recommendation,
                    item.confidence_explanation,
                    round(item.duplicate_confidence, 4),
                    item.source_count,
                    item.report_count,
                    item.merged_sources,
                    item.url,
                    item.ai_reason,
                    item.last_seen_at.isoformat(sep=" ", timespec="seconds"),
                ]
            )

        live_sheet = wb[LIVE_EVENTS_SHEET] if LIVE_EVENTS_SHEET in wb.sheetnames else wb.create_sheet(LIVE_EVENTS_SHEET)
        _ensure_headers(live_sheet, LIVE_EVENT_HEADERS)
        wb.save(path)

    return len(rows)


def append_live_event_to_excel(event: dict[str, str | int | float]) -> None:
    path = Path(settings.excel_path)
    with excel_write_lock:
        wb = _ensure_workbook(path)
        live_sheet = wb[LIVE_EVENTS_SHEET] if LIVE_EVENTS_SHEET in wb.sheetnames else wb.create_sheet(LIVE_EVENTS_SHEET)
        _ensure_headers(live_sheet, LIVE_EVENT_HEADERS)
        live_sheet.append(
            [
                event.get("event_time_utc", ""),
                event.get("event_type", ""),
                event.get("severity", ""),
                event.get("published_at", ""),
                event.get("title", ""),
                event.get("source", ""),
                event.get("language", ""),
                event.get("ai_score", 0),
                event.get("risk_score", 0),
                event.get("crime_type", ""),
                event.get("species", ""),
                event.get("state", ""),
                event.get("district", ""),
                event.get("likely_smuggling_route", ""),
                event.get("enforcement_recommendation", ""),
                event.get("confidence_explanation", ""),
                event.get("source_count", 1),
                event.get("merged_sources", ""),
                event.get("duplicate_confidence", 0),
                event.get("provider", ""),
                event.get("query", ""),
                event.get("url", ""),
            ]
        )
        wb.save(path)
