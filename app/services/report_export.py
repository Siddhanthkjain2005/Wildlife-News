from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO
from statistics import mean

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_csv_bytes(rows: list[dict[str, object]]) -> bytes:
    fields = [
        "date",
        "report_count",
        "risk_score",
        "species",
        "state",
        "district",
        "crime_type",
        "source",
        "confidence",
        "title",
        "two_line_summary",
        "key_intelligence_points",
        "likely_smuggling_route",
        "action_recommendation",
        "confidence_explanation",
        "wpa_schedule",
        "wpa_section",
        "wpa_offence_type",
        "wpa_penalty_class",
        "protected_area_type",
        "enforcement_authority",
        "review_status",
        "reviewed_by",
        "reviewed_at",
        "review_notes",
    ]
    stream = StringIO()
    writer = csv.DictWriter(stream, fieldnames=fields)
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in fields})
    return stream.getvalue().encode("utf-8")


def _date_range(rows: list[dict[str, object]]) -> str:
    values = [str(row.get("date", "")).strip() for row in rows if str(row.get("date", "")).strip()]
    if not values:
        return "-"
    values.sort()
    return f"{values[0][:10]} to {values[-1][:10]}"


def _top_hotspots(rows: list[dict[str, object]], limit: int = 6) -> list[tuple[str, int]]:
    counts: dict[str, int] = {}
    for row in rows:
        state = str(row.get("state", "")).strip() or "Unknown"
        district = str(row.get("district", "")).strip() or "Unknown"
        key = f"{district}, {state}"
        counts[key] = counts.get(key, 0) + 1
    return sorted(counts.items(), key=lambda item: item[1], reverse=True)[:limit]


def _top_recommendations(rows: list[dict[str, object]], limit: int = 6) -> list[tuple[str, int]]:
    counts: dict[str, int] = {}
    for row in rows:
        recommendation = str(row.get("action_recommendation", "")).strip()
        if not recommendation:
            continue
        counts[recommendation] = counts.get(recommendation, 0) + 1
    return sorted(counts.items(), key=lambda item: item[1], reverse=True)[:limit]


def build_excel_bytes(rows: list[dict[str, object]], title: str = "Wildlife Intelligence Report") -> bytes:
    wb = Workbook()
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"
    ws_incidents = wb.create_sheet("Incidents")
    ws_reco = wb.create_sheet("Recommendations")

    heading_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    heading_font = Font(color="FFFFFF", bold=True)

    ws_exec["A1"] = title
    ws_exec["A1"].font = Font(size=16, bold=True)
    ws_exec["A3"] = "Date Range"
    ws_exec["B3"] = _date_range(rows)
    ws_exec["A4"] = "Total Incidents"
    ws_exec["B4"] = len(rows)
    avg_risk = round(mean([float(row.get("risk_score", 0) or 0) for row in rows]), 2) if rows else 0.0
    ws_exec["A5"] = "Average Risk"
    ws_exec["B5"] = avg_risk
    ws_exec["A6"] = "High Risk Incidents (>80)"
    ws_exec["B6"] = sum(1 for row in rows if int(row.get("risk_score", 0) or 0) > 80)
    ws_exec["A8"] = "Top Hotspots"
    ws_exec["A8"].font = Font(bold=True)
    hotspot_row = 9
    for label, count in _top_hotspots(rows):
        ws_exec.cell(row=hotspot_row, column=1, value=label)
        ws_exec.cell(row=hotspot_row, column=2, value=count)
        hotspot_row += 1
    ws_exec.column_dimensions["A"].width = 42
    ws_exec.column_dimensions["B"].width = 20

    incident_fields = [
        "date",
        "report_count",
        "risk_score",
        "species",
        "state",
        "district",
        "crime_type",
        "source",
        "confidence",
        "title",
        "two_line_summary",
        "key_intelligence_points",
        "likely_smuggling_route",
        "action_recommendation",
        "confidence_explanation",
        "wpa_schedule",
        "wpa_section",
        "wpa_offence_type",
        "wpa_penalty_class",
        "protected_area_type",
        "enforcement_authority",
        "review_status",
        "reviewed_by",
        "reviewed_at",
        "review_notes",
    ]
    ws_incidents.append(incident_fields)
    for col_idx in range(1, len(incident_fields) + 1):
        cell = ws_incidents.cell(row=1, column=col_idx)
        cell.fill = heading_fill
        cell.font = heading_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws_incidents.append([row.get(key, "") for key in incident_fields])
    widths = {
        "A": 18,
        "B": 12,
        "C": 10,
        "D": 20,
        "E": 18,
        "F": 20,
        "G": 20,
        "H": 22,
        "I": 12,
        "J": 48,
        "K": 52,
        "L": 54,
        "M": 48,
        "N": 48,
        "O": 52,
        "P": 18,
        "Q": 20,
        "R": 22,
        "S": 18,
        "T": 22,
        "U": 24,
        "V": 15,
        "W": 18,
        "X": 20,
        "Y": 30,
    }
    for column, width in widths.items():
        ws_incidents.column_dimensions[column].width = width
    for row in ws_incidents.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws_reco.append(["Recommendation", "Incident Count"])
    for col_idx in (1, 2):
        cell = ws_reco.cell(row=1, column=col_idx)
        cell.fill = heading_fill
        cell.font = heading_font
    for recommendation, count in _top_recommendations(rows, limit=20):
        ws_reco.append([recommendation, count])
    ws_reco.column_dimensions["A"].width = 90
    ws_reco.column_dimensions["B"].width = 18
    for row in ws_reco.iter_rows(min_row=2):
        row[0].alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_excel_incidents_reports_bytes(
    rows: list[dict[str, object]],
    title: str = "Wildlife Intelligence Incident Export",
) -> bytes:
    wb = Workbook()
    ws_incidents = wb.active
    ws_incidents.title = "Total Incidents"
    ws_reports = wb.create_sheet("Reports Today")

    heading_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    heading_font = Font(color="FFFFFF", bold=True)

    incident_fields = [
        "date",
        "report_count",
        "risk_score",
        "species",
        "state",
        "district",
        "crime_type",
        "source",
        "confidence",
        "title",
        "two_line_summary",
        "key_intelligence_points",
        "likely_smuggling_route",
        "action_recommendation",
        "confidence_explanation",
        "wpa_schedule",
        "wpa_section",
        "wpa_offence_type",
        "wpa_penalty_class",
        "protected_area_type",
        "enforcement_authority",
        "review_status",
        "reviewed_by",
        "reviewed_at",
        "review_notes",
    ]
    ws_incidents.append(incident_fields)
    for col_idx in range(1, len(incident_fields) + 1):
        cell = ws_incidents.cell(row=1, column=col_idx)
        cell.fill = heading_fill
        cell.font = heading_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws_incidents.append([row.get(key, "") for key in incident_fields])
    for column, width in {
        "A": 18,
        "B": 12,
        "C": 10,
        "D": 20,
        "E": 18,
        "F": 20,
        "G": 20,
        "H": 22,
        "I": 12,
        "J": 48,
        "K": 52,
        "L": 54,
        "M": 48,
        "N": 48,
        "O": 52,
        "P": 18,
        "Q": 20,
        "R": 22,
        "S": 18,
        "T": 22,
        "U": 24,
        "V": 15,
        "W": 18,
        "X": 20,
        "Y": 30,
    }.items():
        ws_incidents.column_dimensions[column].width = width
    for row in ws_incidents.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    report_total = int(sum(int(row.get("report_count", 0) or 0) for row in rows))
    ws_reports["A1"] = title
    ws_reports["A1"].font = Font(size=15, bold=True)
    ws_reports["A3"] = "Total Incidents"
    ws_reports["B3"] = len(rows)
    ws_reports["A4"] = "Source Reports Today"
    ws_reports["B4"] = report_total
    ws_reports["A5"] = "Date Range"
    ws_reports["B5"] = _date_range(rows)

    headers = [
        "date",
        "report_count",
        "risk_score",
        "state",
        "district",
        "species",
        "crime_type",
        "source",
        "title",
    ]
    start_row = 7
    for idx, name in enumerate(headers, start=1):
        cell = ws_reports.cell(row=start_row, column=idx, value=name)
        cell.fill = heading_fill
        cell.font = heading_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    reports_rows = sorted(
        rows,
        key=lambda row: (int(row.get("report_count", 0) or 0), str(row.get("date", ""))),
        reverse=True,
    )
    write_row = start_row + 1
    for row in reports_rows:
        ws_reports.cell(row=write_row, column=1, value=row.get("date", ""))
        ws_reports.cell(row=write_row, column=2, value=int(row.get("report_count", 0) or 0))
        ws_reports.cell(row=write_row, column=3, value=int(row.get("risk_score", 0) or 0))
        ws_reports.cell(row=write_row, column=4, value=row.get("state", ""))
        ws_reports.cell(row=write_row, column=5, value=row.get("district", ""))
        ws_reports.cell(row=write_row, column=6, value=row.get("species", ""))
        ws_reports.cell(row=write_row, column=7, value=row.get("crime_type", ""))
        ws_reports.cell(row=write_row, column=8, value=row.get("source", ""))
        ws_reports.cell(row=write_row, column=9, value=row.get("title", ""))
        write_row += 1
    for column, width in {
        "A": 18,
        "B": 14,
        "C": 10,
        "D": 18,
        "E": 20,
        "F": 20,
        "G": 20,
        "H": 24,
        "I": 56,
    }.items():
        ws_reports.column_dimensions[column].width = width
    for row in ws_reports.iter_rows(min_row=start_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_pdf_bytes(rows: list[dict[str, object]], title: str = "Wildlife Intelligence Report") -> bytes:
    buf = BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    y = height - 2 * cm
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(2 * cm, y, title)
    y -= 0.8 * cm
    pdf.setFont("Helvetica", 9)
    pdf.drawString(2 * cm, y, f"Date range: {_date_range(rows)}")
    y -= 0.55 * cm
    pdf.drawString(2 * cm, y, f"Total incidents: {len(rows)}")
    y -= 0.8 * cm
    report_total = int(sum(int(row.get("report_count", 0) or 0) for row in rows))
    pdf.drawString(2 * cm, y, f"Source reports today: {report_total}")
    y -= 0.8 * cm
    high_risk = sum(1 for row in rows if int(row.get("risk_score", 0) or 0) > 80)
    pdf.drawString(2 * cm, y, f"High-risk incidents (>80): {high_risk}")
    y -= 0.8 * cm

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(2 * cm, y, "Top hotspots")
    y -= 0.5 * cm
    pdf.setFont("Helvetica", 9)
    for hotspot, count in _top_hotspots(rows):
        if y < 2.2 * cm:
            pdf.showPage()
            y = height - 2 * cm
            pdf.setFont("Helvetica", 9)
        pdf.drawString(2.3 * cm, y, f"- {hotspot}: {count}")
        y -= 0.42 * cm

    y -= 0.2 * cm
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(2 * cm, y, "Top recommendations")
    y -= 0.5 * cm
    pdf.setFont("Helvetica", 9)
    for recommendation, count in _top_recommendations(rows):
        if y < 2.2 * cm:
            pdf.showPage()
            y = height - 2 * cm
            pdf.setFont("Helvetica", 9)
        pdf.drawString(2.3 * cm, y, f"- ({count}) {recommendation[:95]}")
        y -= 0.42 * cm

    pdf.showPage()
    y = height - 2 * cm
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(2 * cm, y, "High-risk incident details")
    y -= 0.7 * cm
    pdf.setFont("Helvetica", 9)

    for idx, row in enumerate(rows[:220], start=1):
        if int(row.get("risk_score", 0) or 0) <= 80:
            continue
        if y < 2.2 * cm:
            pdf.showPage()
            y = height - 2 * cm
            pdf.setFont("Helvetica", 9)
        line = (
            f"{idx}. {row.get('date', '-')[:16]} | Risk {row.get('risk_score', 0)} | "
            f"{str(row.get('state', '-'))[:15]}/{str(row.get('district', '-'))[:15]} | "
            f"{str(row.get('crime_type', '-'))[:20]}"
        )
        pdf.drawString(2 * cm, y, line)
        y -= 0.45 * cm
        title_line = str(row.get("title", ""))[:105]
        pdf.setFillGray(0.15)
        pdf.drawString(2.4 * cm, y, title_line)
        pdf.setFillGray(0.0)
        y -= 0.42 * cm
        route_line = f"Route: {str(row.get('likely_smuggling_route', '-'))[:95]}"
        pdf.drawString(2.4 * cm, y, route_line)
        y -= 0.38 * cm
        reco_line = f"Action: {str(row.get('action_recommendation', '-'))[:95]}"
        pdf.drawString(2.4 * cm, y, reco_line)
        y -= 0.38 * cm

    pdf.save()
    return buf.getvalue()


def build_dossier_pdf_bytes(row: dict[str, object]) -> bytes:
    buf = BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    # Top border / Seal Area
    y = height - 2 * cm
    pdf.setFillColorRGB(0.06, 0.12, 0.24) # Dark Navy Blue theme
    pdf.rect(1.5 * cm, y - 0.2 * cm, width - 3 * cm, 1.2 * cm, fill=True, stroke=False)
    pdf.setFillColorRGB(1.0, 1.0, 1.0)
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawCentredString(width / 2.0, y + 0.2 * cm, "GOVERNMENT OF INDIA - WILDLIFE CRIME CONTROL BUREAU")
    
    # Reset fill color
    pdf.setFillColorRGB(0.0, 0.0, 0.0)
    y -= 1.2 * cm
    
    # Document Title
    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawString(1.5 * cm, y, "WILDLIFE CRIME INTELLIGENCE DOSSIER")
    pdf.setFont("Helvetica", 9)
    pdf.drawString(width - 4.5 * cm, y, f"Ref: WCCB-ID-{row.get('id', 'N/A')}")
    y -= 0.5 * cm
    
    # Decorative line
    pdf.setStrokeColorRGB(0.7, 0.7, 0.7)
    pdf.setLineWidth(1)
    pdf.line(1.5 * cm, y, width - 1.5 * cm, y)
    y -= 0.6 * cm
    
    # Main Metadata Grid
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(1.5 * cm, y, "INCIDENT INFORMATION DETAILS")
    y -= 0.45 * cm
    
    pdf.setFont("Helvetica", 9)
    metadata = [
        ("Incident Date:", str(row.get("date", ""))[:16]),
        ("Risk Level:", f"{row.get('risk_score', '0')} / 100 (Confidence: {float(row.get('confidence', 0.0) or 0.0):.2f})"),
        ("Target Species:", str(row.get("species", "")) or "—"),
        ("Crime Category:", str(row.get("crime_type", "")) or "—"),
        ("Incident State:", str(row.get("state", "")) or "—"),
        ("Incident District:", str(row.get("district", "")) or "—"),
        ("Involved Suspects:", str(row.get("involved_persons", "")) or "—"),
        ("Intelligence Source:", str(row.get("source", "")) or "—"),
    ]
    
    # Draw metadata in 2 columns
    col_width = (width - 3 * cm) / 2.0
    for idx, (label, val) in enumerate(metadata):
        col = idx % 2
        row_idx = idx // 2
        cur_x = 1.5 * cm + col * col_width
        cur_y = y - row_idx * 0.45 * cm
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(cur_x, cur_y, label)
        pdf.setFont("Helvetica", 9)
        pdf.drawString(cur_x + 3.2 * cm, cur_y, val[:35])
        
    y -= 4 * 0.45 * cm + 0.3 * cm
    pdf.line(1.5 * cm, y, width - 1.5 * cm, y)
    y -= 0.6 * cm
    
    # Title Header Details
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(1.5 * cm, y, "INTELLIGENCE REPORT HEADER")
    y -= 0.45 * cm
    pdf.setFont("Helvetica-Oblique", 9)
    # Wrap title
    title_str = str(row.get("title", ""))
    words = title_str.split()
    lines = []
    curr_line = ""
    for word in words:
        if len(curr_line + " " + word) < 100:
            curr_line += " " + word if curr_line else word
        else:
            lines.append(curr_line)
            curr_line = word
    if curr_line:
        lines.append(curr_line)
    for line in lines[:3]:
        pdf.drawString(1.7 * cm, y, line)
        y -= 0.4 * cm
    y -= 0.2 * cm
    pdf.line(1.5 * cm, y, width - 1.5 * cm, y)
    y -= 0.6 * cm

    # AI Intel Summary Section
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(1.5 * cm, y, "AI INTELLIGENCE REPORT & LEGAL PROVISIONS")
    y -= 0.45 * cm
    
    # Detailed WPA breakdown
    pdf.setFont("Helvetica-Bold", 9)
    wpa_fields = [
        ("WPA Schedule:", str(row.get("wpa_schedule", "")) or "Not Classified"),
        ("WPA Section:", str(row.get("wpa_section", "")) or "Not Classified"),
        ("Offence Category:", str(row.get("wpa_offence_type", "")) or "Not Classified"),
        ("Penalty Class:", str(row.get("wpa_penalty_class", "")) or "Not Classified"),
        ("Protected Area Type:", str(row.get("protected_area_type", "")) or "None"),
        ("Enforcement Agency:", str(row.get("enforcement_authority", "")) or "Forest Dept. / Police"),
    ]
    for idx, (label, val) in enumerate(wpa_fields):
        col = idx % 2
        row_idx = idx // 2
        cur_x = 1.7 * cm + col * col_width
        cur_y = y - row_idx * 0.45 * cm
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(cur_x, cur_y, label)
        pdf.setFont("Helvetica", 9)
        pdf.drawString(cur_x + 3.4 * cm, cur_y, val[:35])
        
    y -= 3 * 0.45 * cm + 0.3 * cm
    pdf.line(1.5 * cm, y, width - 1.5 * cm, y)
    y -= 0.6 * cm
    
    # Intel Details / Route / Recommendations
    details = [
        ("Intelligence Summary Notes:", str(row.get("two_line_summary", "")) or str(row.get("summary", "")) or "No summary notes generated."),
        ("Likely Poaching/Smuggling Route:", str(row.get("likely_smuggling_route", "")) or "Not clear from initial signals."),
        ("Action Enforcement Recommendations:", str(row.get("action_recommendation", "")) or "Review reports and perform physical field verification."),
    ]
    
    for section_title, text in details:
        if y < 3 * cm:
            pdf.showPage()
            y = height - 2 * cm
            
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(1.5 * cm, y, section_title)
        y -= 0.4 * cm
        pdf.setFont("Helvetica", 9)
        
        # Wrap text block
        words = text.split()
        curr_line = ""
        lines = []
        for word in words:
            if len(curr_line + " " + word) < 110:
                curr_line += " " + word if curr_line else word
            else:
                lines.append(curr_line)
                curr_line = word
        if curr_line:
            lines.append(curr_line)
        for line in lines[:4]: # Limit to 4 lines per section
            pdf.drawString(1.7 * cm, y, line)
            y -= 0.38 * cm
        y -= 0.25 * cm
        
    # Signature box at bottom
    if y < 3.5 * cm:
        pdf.showPage()
        y = height - 2 * cm
        
    y = 2.8 * cm
    pdf.line(1.5 * cm, y, width - 1.5 * cm, y)
    y -= 0.6 * cm
    pdf.setFont("Helvetica", 8)
    pdf.drawString(1.5 * cm, y, "REPORT GENERATED BY WILDLIFE CRIME INTELLIGENCE CENTER (WCCB COOPERATION FRAMEWORK).")
    pdf.drawString(1.5 * cm, y - 0.3 * cm, "CLASSIFICATION: FOR OFFICIAL USE ONLY. TIGHT PHYSICAL CONTROL ENFORCED.")
    
    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawString(width - 5 * cm, y, "AUTHORED/STAMPED BY:")
    pdf.setFont("Helvetica-Oblique", 8)
    pdf.drawString(width - 5 * cm, y - 0.4 * cm, "WCCB Duty Intelligence Officer")
    
    pdf.save()
    return buf.getvalue()


def build_bulletin_pdf_bytes(rows: list[dict[str, object]]) -> bytes:
    buf = BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    
    # PAGE 1: EXECUTIVE BRIEFING BULLETIN
    y = height - 1.8 * cm
    
    # Official Header
    pdf.setFont("Helvetica-Bold", 14)
    pdf.setFillColorRGB(0.08, 0.18, 0.36)  # Dark Navy Blue
    pdf.drawCentredString(width / 2.0, y, "WILDLIFE CRIME INTELLIGENCE CENTER")
    y -= 0.45 * cm
    pdf.setFont("Helvetica-Bold", 10)
    pdf.setFillColorRGB(0.72, 0.53, 0.04)  # Gold Accent
    pdf.drawCentredString(width / 2.0, y, "WCCB TACTICAL INTELLIGENCE & SECURITY ADVISORY BULLETIN")
    y -= 0.4 * cm
    pdf.setFont("Helvetica", 8)
    pdf.setFillColorRGB(0.3, 0.3, 0.3)
    pdf.drawCentredString(width / 2.0, y, f"GENERATED ON: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} UTC | FOR OFFICIAL ENFORCEMENT USE ONLY")
    
    y -= 0.4 * cm
    pdf.setStrokeColorRGB(0.08, 0.18, 0.36)
    pdf.setLineWidth(1.5)
    pdf.line(1.5 * cm, y, width - 1.5 * cm, y)
    
    # Metadata Overview block
    y -= 0.8 * cm
    pdf.setFont("Helvetica-Bold", 11)
    pdf.setFillColorRGB(0.08, 0.18, 0.36)
    pdf.drawString(1.5 * cm, y, "I. WEEKLY TASKFORCE INTEL SUMMARY")
    
    y -= 0.5 * cm
    pdf.setFont("Helvetica", 9.5)
    pdf.setFillColorRGB(0.1, 0.1, 0.1)
    
    total_count = len(rows)
    high_risk_count = sum(1 for r in rows if int(r.get("risk_score", 0) or 0) > 80)
    
    # Get distinct states affected
    states = {str(r.get("state", "")).strip() for r in rows if r.get("state")}
    states.discard("")
    species = {str(r.get("species", "")).strip() for r in rows if r.get("species")}
    species.discard("")
    
    summary_text = (
        f"During the past 7 days, a total of {total_count} active wildlife crime incidents have been intercepted "
        f"and analyzed by the hybrid AI pipeline. Out of these, {high_risk_count} cases are classified as HIGH-RISK alerts "
        f"requiring immediate field intervention. Poaching activity was detected across {len(states)} states "
        f"impacting {len(species)} federally protected species."
    )
    
    # Wrap summary text
    words = summary_text.split()
    curr_line = ""
    lines = []
    for word in words:
        if len(curr_line + " " + word) < 110:
            curr_line += " " + word if curr_line else word
        else:
            lines.append(curr_line)
            curr_line = word
    if curr_line:
        lines.append(curr_line)
    for line in lines:
        pdf.drawString(1.5 * cm, y, line)
        y -= 0.42 * cm
        
    # Hotspot analysis grid
    y -= 0.4 * cm
    pdf.line(1.5 * cm, y, width - 1.5 * cm, y)
    y -= 0.6 * cm
    
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(1.5 * cm, y, "II. TACTICAL COMPLIANCE & SPECIES IMPACT")
    y -= 0.5 * cm
    
    pdf.setFont("Helvetica", 9)
    hotspots = _top_hotspots(rows, limit=5)
    pdf.drawString(1.5 * cm, y, "TOP AFFECTED DISTRICTS (HOTSPOTS):")
    y -= 0.4 * cm
    for place, count in hotspots:
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(1.8 * cm, y, f"• {place}:")
        pdf.setFont("Helvetica", 9)
        pdf.drawString(5.5 * cm, y, f"{count} threat reports compiled")
        y -= 0.38 * cm
        
    y -= 0.4 * cm
    pdf.setFont("Helvetica", 9)
    pdf.drawString(1.5 * cm, y, "MOST IMPACTED SCHEDULED SPECIES:")
    y -= 0.4 * cm
    
    # Species count
    sp_counts = {}
    for r in rows:
        sp = str(r.get("species", "")).strip()
        if sp and sp != "—":
            sp_counts[sp] = sp_counts.get(sp, 0) + 1
    sorted_sp = sorted(sp_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    for sp_name, count in sorted_sp:
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(1.8 * cm, y, f"• {sp_name.title()}:")
        pdf.setFont("Helvetica", 9)
        pdf.drawString(5.5 * cm, y, f"{count} cases registered")
        y -= 0.38 * cm
        
    # Legal Warning Block
    y -= 0.5 * cm
    pdf.setStrokeColorRGB(0.72, 0.53, 0.04)
    pdf.setFillColorRGB(0.99, 0.98, 0.94)
    pdf.rect(1.5 * cm, y - 1.6 * cm, width - 3 * cm, 1.8 * cm, fill=True, stroke=True)
    
    pdf.setFillColorRGB(0.08, 0.18, 0.36)
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(1.8 * cm, y - 0.3 * cm, "OFFICIAL STATUTORY NOTICE - WILDLIFE PROTECTION ACT (WPA), 1972")
    pdf.setFont("Helvetica-Oblique", 8.5)
    pdf.drawString(1.8 * cm, y - 0.7 * cm, "All listed Schedule I and Schedule II species represent highest conservation priority categories.")
    pdf.drawString(1.8 * cm, y - 1.0 * cm, "Any trade, hunting, or possession thereof carries strict non-bailable imprisonment from 3 to 7 years.")
    pdf.drawString(1.8 * cm, y - 1.3 * cm, "Range forest officers are authorized to arrest suspects and seize contraband under Section 50 of the Act.")
    
    y -= 2.2 * cm
    pdf.drawString(1.5 * cm, y, "CLASSIFICATION: DEPARTMENTAL USE ONLY | STAMP & SIGNATURE REQUIREMENT")
    pdf.line(1.5 * cm, y - 0.2 * cm, width - 1.5 * cm, y - 0.2 * cm)
    
    y -= 1.2 * cm
    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawString(1.5 * cm, y, "COMPILED BY:")
    pdf.drawString(width / 2.0, y, "VERIFIED BY:")
    pdf.drawString(width - 5 * cm, y, "STAMP OF AUTHORITY:")
    
    pdf.setFont("Helvetica-Oblique", 8)
    pdf.drawString(1.5 * cm, y - 0.4 * cm, "WTI Crime Analyst")
    pdf.drawString(width / 2.0, y - 0.4 * cm, "WCCB Duty Coordinator")
    
    # PAGE 2: INCIDENT CHRONOLOGICAL BRIEFING
    pdf.showPage()
    y = height - 1.8 * cm
    
    pdf.setFont("Helvetica-Bold", 12)
    pdf.setFillColorRGB(0.08, 0.18, 0.36)
    pdf.drawString(1.5 * cm, y, "III. CHRONOLOGICAL TACTICAL INCIDENT CHANNELS")
    y -= 0.3 * cm
    pdf.setStrokeColorRGB(0.08, 0.18, 0.36)
    pdf.setLineWidth(1.0)
    pdf.line(1.5 * cm, y, width - 1.5 * cm, y)
    y -= 0.6 * cm
    
    # Print up to 10 latest incidents
    pdf.setFont("Helvetica", 9)
    for idx, r in enumerate(rows[:9]):
        if y < 2.5 * cm:
            pdf.showPage()
            y = height - 1.8 * cm
            
        pdf.setFont("Helvetica-Bold", 9.5)
        pdf.setFillColorRGB(0.08, 0.18, 0.36)
        date_str = str(r.get("date", ""))[:10]
        pdf.drawString(1.5 * cm, y, f"[{date_str}] CASE ID: {r.get('id', '—')} | {r.get('species', 'Unknown Species')} | Score: {r.get('risk_score', '0')}")
        y -= 0.38 * cm
        
        pdf.setFont("Helvetica-Bold", 9.5)
        pdf.setFillColorRGB(0.72, 0.53, 0.04)
        pdf.drawString(1.8 * cm, y, "Location:")
        pdf.setFont("Helvetica", 9)
        pdf.setFillColorRGB(0.1, 0.1, 0.1)
        pdf.drawString(3.3 * cm, y, f"{r.get('district', '—')}, {r.get('state', '—')} (Source: {r.get('source', '—')})")
        y -= 0.38 * cm
        
        pdf.setFont("Helvetica-Bold", 9.5)
        pdf.setFillColorRGB(0.72, 0.53, 0.04)
        pdf.drawString(1.8 * cm, y, "Summary:")
        pdf.setFont("Helvetica", 9)
        pdf.setFillColorRGB(0.1, 0.1, 0.1)
        
        title_text = str(r.get("title", ""))
        words = title_text.split()
        curr_line = ""
        lines = []
        for word in words:
            if len(curr_line + " " + word) < 95:
                curr_line += " " + word if curr_line else word
            else:
                lines.append(curr_line)
                curr_line = word
        if curr_line:
            lines.append(curr_line)
        for line in lines[:2]:
            pdf.drawString(3.3 * cm, y, line)
            y -= 0.35 * cm
            
        y -= 0.25 * cm
        pdf.setStrokeColorRGB(0.9, 0.9, 0.9)
        pdf.line(1.7 * cm, y, width - 1.7 * cm, y)
        y -= 0.45 * cm
        
    pdf.save()
    return buf.getvalue()
